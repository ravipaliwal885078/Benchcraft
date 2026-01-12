"""
Allocation Reports Routes
Generates Internal Allocation Reports and Requisition Reports (Client-facing)
Supports both project-level and overall reports
"""
from flask import Blueprint, request, jsonify, send_file
from tools.sql_db import SQLDatabaseTool
from models import (
    Employee, Allocation, Project, EmployeeSkill, AllocationFinancial,
    RiskRegister, RiskType
)
from datetime import date, datetime, timedelta
from sqlalchemy import and_, or_, func
from io import BytesIO
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not installed. Excel export will not be available.")

bp = Blueprint('allocation_reports', __name__)
db_tool = SQLDatabaseTool()


def get_primary_skills(employee, session, limit=3):
    """Get top N primary skills for an employee"""
    skills = session.query(EmployeeSkill).filter(
        EmployeeSkill.emp_id == employee.id
    ).order_by(EmployeeSkill.proficiency.desc()).limit(limit).all()
    return ', '.join([skill.skill_name for skill in skills])


def calculate_monthly_hours(allocation_percentage):
    """Calculate monthly hours: (160 × allocation_percentage) / 100"""
    return (160 * allocation_percentage) / 100


def calculate_billable_hours(allocation_percentage, billable_percentage):
    """Calculate billable hours: (160 × allocation% × billable%) / 10000"""
    return (160 * allocation_percentage * billable_percentage) / 10000


def get_utilization_status(allocation_percentage):
    """Determine utilization status"""
    if allocation_percentage < 80:
        return "Under-utilized"
    elif allocation_percentage <= 100:
        return "Optimal"
    else:
        return "Over-allocated"


def get_project_status(allocation, today=None):
    """Determine project status based on dates"""
    if today is None:
        today = date.today()
    
    if allocation.end_date and allocation.end_date < today:
        return "Delayed"
    elif allocation.end_date and allocation.end_date <= today + timedelta(days=7):
        return "On-Track"
    else:
        return "Ahead"


def filter_client_risks(risks, session):
    """Filter risks to only show client-relevant ones (exclude internal HR risks)"""
    client_relevant_risks = []
    for risk in risks:
        # Exclude internal HR risks like resignation
        if risk.risk_type and risk.risk_type.upper() not in ['RESIGNATION', 'NOTICE_PERIOD']:
            client_relevant_risks.append(risk)
    return client_relevant_risks


@bp.route('/generate', methods=['GET'])
def generate_allocation_report():
    """
    Generate Allocation Report (JSON format)
    Query params:
    - report_type: 'internal' or 'requisition' (default: 'internal')
    - level: 'project' or 'overall' (default: 'overall')
    - project_id: Required if level='project'
    - start_date: Report period start (YYYY-MM-DD)
    - end_date: Report period end (YYYY-MM-DD)
    """
    session = db_tool.Session()
    try:
        report_type = request.args.get('report_type', 'internal').lower()
        level = request.args.get('level', 'overall').lower()
        project_id = request.args.get('project_id', type=int)
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        
        # Validate parameters
        if level == 'project' and not project_id:
            return jsonify({'error': 'project_id is required for project-level reports'}), 400
        
        # Parse dates
        today = date.today()
        if start_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        else:
            start_date = today
        
        if end_date_str:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        else:
            end_date = today + timedelta(days=30)
        
        # Get project if project-level report
        project = None
        if level == 'project' and project_id:
            project = session.query(Project).filter(Project.id == project_id).first()
            if not project:
                return jsonify({'error': 'Project not found'}), 404
        
        # Get allocations
        # Explicitly specify join conditions to avoid ambiguity (Allocation has multiple FKs to Employee)
        query = session.query(Allocation).join(Employee, Allocation.emp_id == Employee.id).join(Project, Allocation.proj_id == Project.id)
        
        # Filter by project if project-level
        if level == 'project' and project_id:
            query = query.filter(Allocation.proj_id == project_id)
        
        # Filter by date range
        query = query.filter(
            and_(
                Allocation.start_date <= end_date,
                or_(
                    Allocation.end_date.is_(None),
                    Allocation.end_date >= start_date
                )
            )
        )
        
        # Exclude trainees from requisition reports
        if report_type == 'requisition':
            query = query.filter(
                or_(
                    Allocation.is_trainee.is_(None),
                    Allocation.is_trainee == False
                )
            )
        
        allocations = query.all()
        
        # Build report data
        report_data = []
        for alloc in allocations:
            employee = alloc.employee
            proj = alloc.project
            
            # Skip if employee or project not found
            if not employee or not proj:
                continue
            
            # Get allocation percentages
            allocation_pct = getattr(alloc, 'allocation_percentage', 100) or 100
            billable_pct = getattr(alloc, 'billable_percentage', 100) or 100
            internal_pct = getattr(alloc, 'internal_allocation_percentage', allocation_pct) or allocation_pct
            
            # Calculate metrics
            monthly_hours = calculate_monthly_hours(allocation_pct)
            billable_hours = calculate_billable_hours(allocation_pct, billable_pct)
            hourly_rate = alloc.billing_rate or 0
            monthly_amount = hourly_rate * billable_hours
            
            # Get period revenue from allocation_financials if available
            period_revenue = None
            if alloc.financials:
                period_revenue = getattr(alloc.financials, 'estimated_revenue', None)
            
            # Get primary skills
            primary_skills = get_primary_skills(employee, session)
            
            # Build resource detail
            resource_detail = {
                'id': employee.id,
                'employee_id': employee.id,  # For backward compatibility
                'employee_name': f"{employee.first_name} {employee.last_name}",
                'first_name': employee.first_name,
                'last_name': employee.last_name,
                'email': employee.email,
                'role': employee.role_level or 'N/A',
                'role_level': employee.role_level,
                'primary_skills': primary_skills,
                'start_date': alloc.start_date.isoformat() if alloc.start_date else None,
                'end_date': alloc.end_date.isoformat() if alloc.end_date else 'Ongoing',
                'allocation_percentage': allocation_pct,
                'billable_percentage': billable_pct,
                'monthly_hours': round(monthly_hours, 2),
                'billable_hours': round(billable_hours, 2),
                'hourly_rate': hourly_rate,
                'monthly_amount': round(monthly_amount, 2),
                'period_revenue': round(period_revenue, 2) if period_revenue else None,
                'utilization': get_utilization_status(allocation_pct),
                'status': get_project_status(alloc, today),
                'project_id': proj.id,
                'project_name': proj.project_name,
                'client_name': proj.client_name,
                'project_code': getattr(proj, 'project_code', None),
                'currency': getattr(proj, 'billing_currency', 'USD')
            }
            
            # Add internal allocation percentage only for internal reports
            if report_type == 'internal':
                resource_detail['internal_allocation_percentage'] = internal_pct
            
            report_data.append(resource_detail)
        
        # Calculate summary
        total_resources = len(report_data)
        total_monthly_hours = sum(r['monthly_hours'] for r in report_data)
        total_billable_hours = sum(r['billable_hours'] for r in report_data)
        total_monthly_amount = sum(r['monthly_amount'] for r in report_data)
        avg_allocation = sum(r['allocation_percentage'] for r in report_data) / total_resources if total_resources > 0 else 0
        avg_billable = sum(r['billable_percentage'] for r in report_data) / total_resources if total_resources > 0 else 0
        
        # Build response
        response = {
            'report_type': report_type,
            'level': level,
            'report_date': today.isoformat(),
            'reporting_period': {
                'start_date': start_date.isoformat(),
                'end_date': end_date.isoformat()
            },
            'project': {
                'id': project.id if project else None,
                'name': project.project_name if project else None,
                'code': getattr(project, 'project_code', None) if project else None,
                'client_name': project.client_name if project else None,
                'billing_currency': getattr(project, 'billing_currency', 'USD') if project else 'USD'
            } if project else None,
            'resources': report_data,
            'summary': {
                'total_resources': total_resources,
                'total_monthly_hours': round(total_monthly_hours, 2),
                'total_billable_hours': round(total_billable_hours, 2),
                'total_monthly_amount': round(total_monthly_amount, 2),
                'average_allocation_percentage': round(avg_allocation, 2),
                'average_billable_percentage': round(avg_billable, 2)
            }
        }
        
        return jsonify(response)
        
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"Error generating allocation report: {e}")
        print(error_trace)
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()


@bp.route('/export-excel', methods=['GET'])
def export_allocation_report_excel():
    """
    Export Allocation Report as Excel file
    Query params: Same as /generate endpoint
    """
    if not OPENPYXL_AVAILABLE:
        return jsonify({'error': 'openpyxl library is required for Excel export. Please install it: pip install openpyxl'}), 500
    
    session = db_tool.Session()
    try:
        report_type = request.args.get('report_type', 'internal').lower()
        level = request.args.get('level', 'overall').lower()
        project_id = request.args.get('project_id', type=int)
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        
        # Validate parameters
        if level == 'project' and not project_id:
            return jsonify({'error': 'project_id is required for project-level reports'}), 400
        
        # Parse dates
        today = date.today()
        if start_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        else:
            start_date = today
        
        if end_date_str:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        else:
            end_date = today + timedelta(days=30)
        
        # Get project if project-level report
        project = None
        if level == 'project' and project_id:
            project = session.query(Project).filter(Project.id == project_id).first()
            if not project:
                return jsonify({'error': 'Project not found'}), 404
        
        # Get allocations (same logic as generate endpoint)
        # Explicitly specify join conditions to avoid ambiguity (Allocation has multiple FKs to Employee)
        query = session.query(Allocation).join(Employee, Allocation.emp_id == Employee.id).join(Project, Allocation.proj_id == Project.id)
        
        if level == 'project' and project_id:
            query = query.filter(Allocation.proj_id == project_id)
        
        query = query.filter(
            and_(
                Allocation.start_date <= end_date,
                or_(
                    Allocation.end_date.is_(None),
                    Allocation.end_date >= start_date
                )
            )
        )
        
        if report_type == 'requisition':
            query = query.filter(
                or_(
                    Allocation.is_trainee.is_(None),
                    Allocation.is_trainee == False
                )
            )
        
        allocations = query.all()
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resource Allocation Report"
        
        # Styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=14)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Section A: Report Header (Rows 1-6)
        row = 1
        
        # Client Name
        ws['A1'] = 'Client Name'
        ws['B1'] = project.client_name if project else 'All Projects'
        ws['A1'].font = Font(bold=True)
        
        # Project Name
        ws['A2'] = 'Project Name'
        ws['B2'] = project.project_name if project else 'Overall Allocation'
        ws['A2'].font = Font(bold=True)
        
        # Project Code
        if project and hasattr(project, 'project_code'):
            ws['A3'] = 'Project Code'
            ws['B3'] = project.project_code or 'N/A'
            ws['A3'].font = Font(bold=True)
        
        # Reporting Period
        ws['A4'] = 'Reporting Period'
        ws['B4'] = f"{start_date.strftime('%Y-%m-%d')} - {end_date.strftime('%Y-%m-%d')}"
        ws['A4'].font = Font(bold=True)
        
        # Report Date
        ws['A5'] = 'Report Date'
        ws['B5'] = today.strftime('%Y-%m-%d')
        ws['A5'].font = Font(bold=True)
        
        # Currency
        ws['A6'] = 'Currency'
        currency = getattr(project, 'billing_currency', 'USD') if project else 'USD'
        ws['B6'] = currency
        ws['A6'].font = Font(bold=True)
        
        # Section B: Resource Details (Starting Row 8)
        row = 8
        
        # Column headers
        headers = [
            'Sr. No.', 'Employee Name', 'Role', 'Primary Skills', 'Start Date', 'End Date',
            'Allocation %', 'Billable %', 'Monthly Hours', 'Billable Hours',
            'Hourly Rate', 'Monthly Amount', 'Period Revenue', 'Utilization', 'Status'
        ]
        
        # Add internal allocation column only for internal reports
        if report_type == 'internal':
            headers.insert(7, 'Internal Allocation %')
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Data rows
        row += 1
        sr_no = 1
        
        for alloc in allocations:
            employee = alloc.employee
            proj = alloc.project
            
            if not employee or not proj:
                continue
            
            # Get allocation percentages
            allocation_pct = getattr(alloc, 'allocation_percentage', 100) or 100
            billable_pct = getattr(alloc, 'billable_percentage', 100) or 100
            internal_pct = getattr(alloc, 'internal_allocation_percentage', allocation_pct) or allocation_pct
            
            # Calculate metrics
            monthly_hours = calculate_monthly_hours(allocation_pct)
            billable_hours = calculate_billable_hours(allocation_pct, billable_pct)
            hourly_rate = alloc.billing_rate or 0
            monthly_amount = hourly_rate * billable_hours
            
            # Get period revenue
            period_revenue = None
            if alloc.financials:
                period_revenue = getattr(alloc.financials, 'estimated_revenue', None)
            
            # Get primary skills
            primary_skills = get_primary_skills(employee, session)
            
            # Build row data
            # Note: Percentages need to be stored as decimals (0.5 for 50%) for Excel formatting
            row_data = [
                sr_no,
                f"{employee.first_name} {employee.last_name}",
                employee.role_level or 'N/A',
                primary_skills,
                alloc.start_date.strftime('%Y-%m-%d') if alloc.start_date else '',
                alloc.end_date.strftime('%Y-%m-%d') if alloc.end_date else 'Ongoing',
                allocation_pct / 100.0,  # Convert to decimal for Excel percentage format
                billable_pct / 100.0     # Convert to decimal for Excel percentage format
            ]
            
            # Add internal allocation only for internal reports
            if report_type == 'internal':
                row_data.append(internal_pct / 100.0)  # Convert to decimal
            
            row_data.extend([
                round(monthly_hours, 2),
                round(billable_hours, 2),
                hourly_rate,
                round(monthly_amount, 2),
                round(period_revenue, 2) if period_revenue else '',
                get_utilization_status(allocation_pct),
                get_project_status(alloc, today)
            ])
            
            # Write row
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # Format numbers based on column position
                # Columns: 1=SrNo, 2=Name, 3=Role, 4=Skills, 5=StartDate, 6=EndDate, 
                #          7=Alloc%, 8=Billable%, [9=Internal% if internal], 10=MonthlyHours, 11=BillableHours,
                #          12=HourlyRate, 13=MonthlyAmount, 14=PeriodRevenue, 15=Utilization, 16=Status
                
                if col_idx == 7 or col_idx == 8:  # Allocation % and Billable %
                    cell.number_format = '0%'
                elif report_type == 'internal' and col_idx == 9:  # Internal Allocation %
                    cell.number_format = '0%'
                elif (report_type == 'internal' and col_idx in [10, 11]) or (report_type == 'requisition' and col_idx in [9, 10]):  # Hours columns
                    cell.number_format = '0.00'
                elif (report_type == 'internal' and col_idx in [12, 13, 14]) or (report_type == 'requisition' and col_idx in [11, 12, 13]):  # Currency columns
                    cell.number_format = f'"{currency} "#,##0.00'
            
            row += 1
            sr_no += 1
        
        # Section C: Summary (Bottom rows)
        summary_row = row + 2
        ws.cell(row=summary_row, column=1, value='Summary').font = title_font
        
        summary_row += 1
        summary_data = [
            ['Total Resources', len(allocations)],
            ['Total Monthly Hours', round(sum(calculate_monthly_hours(getattr(a, 'allocation_percentage', 100) or 100) for a in allocations), 2)],
            ['Total Billable Hours', round(sum(calculate_billable_hours(
                getattr(a, 'allocation_percentage', 100) or 100,
                getattr(a, 'billable_percentage', 100) or 100
            ) for a in allocations), 2)],
            ['Total Monthly Amount', round(sum(
                (getattr(a, 'billing_rate', 0) or 0) * calculate_billable_hours(
                    getattr(a, 'allocation_percentage', 100) or 100,
                    getattr(a, 'billable_percentage', 100) or 100
                ) for a in allocations
            ), 2)]
        ]
        
        if allocations:
            avg_alloc = sum(getattr(a, 'allocation_percentage', 100) or 100 for a in allocations) / len(allocations)
            avg_bill = sum(getattr(a, 'billable_percentage', 100) or 100 for a in allocations) / len(allocations)
            summary_data.extend([
                ['Average Allocation %', avg_alloc / 100.0],  # Convert to decimal for Excel
                ['Average Billable %', avg_bill / 100.0]        # Convert to decimal for Excel
            ])
        
        for idx, (label, value) in enumerate(summary_data, start=summary_row):
            ws.cell(row=idx, column=1, value=label).font = Font(bold=True)
            ws.cell(row=idx, column=2, value=value)
            if 'Amount' in label:
                ws.cell(row=idx, column=2).number_format = f'"{currency} "#,##0.00'
            elif '%' in label:
                ws.cell(row=idx, column=2).number_format = '0.00%'  # Excel percentage format
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        filename = f"allocation_report_{report_type}_{level}"
        if project:
            filename += f"_{project.id}"
        filename += f"_{today.strftime('%Y%m%d')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"Error exporting allocation report: {e}")
        print(error_trace)
        return jsonify({'error': str(e)}), 500
    finally:
        session.close()
